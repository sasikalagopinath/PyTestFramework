import openpyxl

def getRowCount(path, sheet):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return sheet.max_row

def getColumnCount(path, sheet):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return sheet.max_column

def readData(path, sheet, r, c):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return sheet.cell(row=r, column=c).value
